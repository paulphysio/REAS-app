from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, get_object_or_404, redirect
from .models import activity, askedQuestions
import pandas as pd
import openpyxl as op
from auth_sys.models import profile
from django.contrib import messages
from django.views.generic import DeleteView
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import AnonymousUser
from .forms import ActivityForm
from django.utils import timezone

def handler404(request, exception):
    return render(request, '404.html', status=404)

def handler500(request):
    return render(request, '500.html', status=500)

def questionAskedView(request):
    if request.method=='POST':
        
        question = request.POST['question']
        askedQuestions(question=question).save()
        return redirect('about')
    return render(request, "about.html")

# class deleteView(DeleteView):
#     model = activity
#     template_name = "delete_file.html"
#     success_url='/'
    
    
    
# delete view for details
def deleteView(request, pk):
    # dictionary for initial data with
    # field names as keys
    
 
    # fetch the object related to passed id
    obj = get_object_or_404(activity, id = pk)
    act_id = obj.id
    
 
    if request.method =="POST":
        # delete object
        obj.delete()
        # after deleting redirect to
        # home page
        return HttpResponseRedirect("/home")
    context ={}
    context['object'] = 28
    return render(request, "base.html", context)
   
# Create your views here.
@login_required
def homeView(request):
    activity_list = activity.objects.all()
    first = True
    if activity_list:
        for list in activity_list:
            if first:
                list = list
                first = False
        context = {
            "activity_list":activity_list,
            "activity_detail":list,
        
        }
    else:
        context = {
        }
    return render(request, "home.html", context)
def aboutView(request):
    return render(request, "about.html")
def aboutUsView(request):
    return render(request, "about_us.html")
def guidesView(request):
    return render(request, "guides.html")
@login_required
def detailView(request, pk):
    activity_detail = activity.objects.get(id=pk)
    sheet = activity_detail.file_uploaded
    #read excel file or data file
    df = pd.read_excel(sheet)
    #load excel sheet as work book
    wb = op.load_workbook(sheet)
    #activate work book
    wa = wb.active
    
    
    lister = []
    reallist = []
    #for every row starting from index 1
    for i in range(1, wa.max_row+1):
        list=[]
        #for every column starting from index 1
        for j in range(1, wa.max_column+1):
            cell_obj = wa.cell(row=i, column=j)
            x = wa.cell(row=i, column=j).value
            list.append(x)
            #if the last column is reached
            if j==(wa.max_column):
                reallist.append(list)
    
    context = {
        "activity_detail":activity_detail,
        "sheet":sheet,
        "df":df,
        "lister":lister,
        "reallist":reallist
    }
    return render(request, "file_details.html", context)

@login_required
def activityFormView(request):
    file_form = ActivityForm
    if request.method=='POST':
        user = request.user
        file_uploaded = request.FILES['uploaded_file']
        result_level = request.POST['result_level']
        Department = request.POST['Department']
        session = request.POST['session']
        semester = request.POST['semester']
        
        #verifying that the file must be an excel sheet
        
      
        if  file_uploaded.name.endswith('xlsx'):

            # Load the Excel file
            wb = op.load_workbook(file_uploaded)

            # Select the active worksheet
            ws = wb.active

            # Iterate over all cells in the worksheet
            for row in ws.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        # Get the hyperlink value
                        value = cell.value
                        # Remove the hyperlink
                        cell.hyperlink = None
                        # Set the cell value to the hyperlink value
                        cell.value = value

            # Save the changes to the workbook
            wb.save(file_uploaded)
            form = activity.objects.create(user = user, file_uploaded = file_uploaded, result_level=result_level, Department=Department, session=session, semester=semester)
            file_path=form.file_uploaded.path
            form.delete_after_delay()
            
        else:
            messages.error(request, f'wrong format')   
        return HttpResponseRedirect('/home')
        
    return render(request, "home.html", {"file_form": file_form} )